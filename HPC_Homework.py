import xlrd, xlwt
import os
import logging
from collections import Counter
import openpyxl

logging.basicConfig(
    #日志级别
    level=logging.DEBUG,
    #日志格式
    format='%(asctime)s %(filename)s[line:%(lineno)d] %(levelname)s %(message)s',
    #日志存放目录
    filename='log.txt',
    #打开日志文件的方式
    filemode="a"
    )


class Split_Excel:
    def __init__(self, FileName):
        self.id = []
        self.id_n =[]
        self.flag = 1
        try:
            self.workbook = xlrd.open_workbook(FileName, formatting_info=False)
            #创建Excel工作簿操作方法，否则文件不存在，任务终止
        except:
            logging.error("%s 文件不存在"%(FileName))
            self.flag = 0

    def openSheet (self, sheetName):
        try:
            self.sheet = self.workbook.sheet_by_name(self.workbook.sheet_names()[sheetName - 1])
            # 打开Sheet表，否则sheetn(n>=2)不存在，开始分割任务
            return True
        except:
            logging.error("Sheet%s 表不存在,开始分割。"%(str(sheetName)))
            return False

    def JudgeIfRepeated (self, sheetName):
        # 判断是否满足id项没有重复
        set_id = set(self.id)
        if len(set_id) !=len(self.id):
            logging.error("sheet%d的id出现不唯一值，任务停止。其中出现次数如下："%(sheetName))
            idlst = self.id
            idlst.remove('id')
            logging.error(Counter(idlst))
            return False
        return True

    def JudgeIfMoreId1 (self, sheetName):
        # 判断是否满足idn(n>=2)的数据是id的子集
        repeated = []
        for element in self.id_n:
            if element not in self.id:
                repeated.append(element)
        if len(repeated) > 0:
            logging.error("sheet%d的出现了不是sheet1的id值，其具体为"%(sheetName))
            logging.error(repeated)
            return False
        return True

    def WriteInExcel (self, data, ID_Name):
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            for r in range(2):
                for c in range(len(data[0])):
                    try:
                        ws.cell(r + 1, c + 1, data[r][c])
                        # 字典的数据维度有可能不相等，所以填空处理
                    except:
                        ws.cell(r + 1, c + 1, '')
            wb.save('test-%s.xlsx'%(str(ID_Name))) # id数据名可能是数字，文字，字符的任何东西
            return True
        except:
            logging.error("id%s写入失败"%(str(ID_Name)))
            return False

    def MainWorkFlow (self):
        if self.flag == 1:
            cnt = 1
            DataName = []
            Data = []
            DataDict = {}
            flag = 1
            while(True):
                if self.openSheet(cnt):
                    if cnt == 1:
                        DataName = DataName + self.sheet.row_values(0)
                        # 存储数据项名称
                        self.id = self.sheet.col_values(0)
                        if not self.JudgeIfRepeated(cnt):
                            flag = 0
                            break
                        Data = Data + self.sheet.col_values(0)[1:]
                        for i in range(len(self.sheet.col_values(0)) - 1):
                            DataDict[Data[i]] = self.sheet.row_values(i + 1)[1:]
                            # 第一次字典中存储sheet1中的数据 
                    else:
                        DataName = DataName + self.sheet.row_values(0)[1:]
                        self.id_n = self.sheet.col_values(0)
                        if not self.JudgeIfMoreId1(cnt):
                            flag = 0
                            break
                        for i in range(len(self.sheet.col_values(0)) - 1):
                            DataDict[self.sheet.col_values(0)[i + 1]] += self.sheet.row_values(i + 1)[1:]
                            # 每次向字典对应项加入新的数据
                        for i in self.id:
                            if i not in self.id_n:
                                DataDict[i] += [' '] * (self.sheet.row_len(0) - 1)
                                # 这步处理主要是解决：有的数据出现在了sheet1和sheet3中，而sheet2中未出现，
                                # 要防止sheet3的数据填充在sheet2的表项中
                else:
                    break
                cnt += 1
            if flag == 1:
                # 如果以上错误皆未发生，说明分割完成，调用写入函数写入Excel中
                for i in self.id:
                    data = []
                    data.append(DataName)
                    try:
                        data.append([i] + DataDict[i])
                    except:
                        continue
                    if not self.WriteInExcel(data, i):
                        break

if __name__ == "__main__":
    work = Split_Excel('test.xlsx')
    work.MainWorkFlow()